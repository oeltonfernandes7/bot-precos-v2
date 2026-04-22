import streamlit as st
import pandas as pd
import plotly.express as px
import os
import json
from datetime import datetime
import time
import random
from io import BytesIO

# Importar funções do bot
from bot_precos import (
    buscar_produto,
    carregar_estrategia,
    carregar_fontes,
    salvar_fontes,
    configurar_logging
)

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Bot de Preços v2.0",
    page_icon="💊",
    layout="wide"
)

# ─────────────────────────────────────────────
# FUNÇÕES UTILITÁRIAS — ARQUIVOS E FORMATAÇÃO
# ─────────────────────────────────────────────

def salvar_csv_historico(df, nome_arquivo=None):
    """Salva DataFrame em CSV no diretório historico."""
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"historico/pesquisa_{timestamp}.csv"
    os.makedirs("historico", exist_ok=True)
    df.to_csv(nome_arquivo, index=False, sep=";", encoding="utf-8-sig")
    return nome_arquivo


def carregar_historico():
    """Carrega todos os CSVs do diretório historico."""
    if not os.path.exists("historico"):
        return pd.DataFrame()
    dfs = []
    for arquivo in os.listdir("historico"):
        if arquivo.endswith(".csv"):
            try:
                df = pd.read_csv(f"historico/{arquivo}", sep=";", encoding="utf-8-sig")
                df["arquivo_origem"] = arquivo
                dfs.append(df)
            except Exception as e:
                st.warning(f"Erro ao carregar {arquivo}: {e}")
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    return pd.DataFrame()


def _carregar_fontes_json():
    """Lê fontes.json com encoding correto e garante a chave 'pendentes'."""
    try:
        with open("fontes.json", "r", encoding="utf-8") as f:
            dados = json.load(f)
        if "pendentes" not in dados:
            dados["pendentes"] = []
        return dados
    except Exception:
        return {"fontes": [], "pendentes": []}


def _atualizar_status_fonte(nome, novo_status):
    """Muda o status de uma fonte no fontes.json e salva."""
    dados = _carregar_fontes_json()
    for fonte in dados.get("fontes", []):
        if fonte["nome"] == nome:
            fonte["status"] = novo_status
    salvar_fontes(dados.get("fontes", []), dados.get("pendentes", []))


def _sync_fonte_para_tab1(coluna_csv):
    """Callback: propaga mudança de checkbox do Tab3 para o Tab1."""
    st.session_state[f"tab1_fonte_{coluna_csv}"] = \
        st.session_state[f"tab3_fonte_{coluna_csv}"]


# ─────────────────────────────────────────────
# FUNÇÕES UTILITÁRIAS — GERAÇÃO DE EXCEL
# ─────────────────────────────────────────────

def formatar_preco_excel(writer, sheet_name, df):
    """Formata colunas de preço e cabeçalho no Excel."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    fmt_preco = workbook.add_format({'num_format': 'R$ #,##0.00', 'align': 'right'})
    fmt_header = workbook.add_format({
        'bold': True, 'bg_color': '#2E75B6',
        'font_color': 'white', 'align': 'center'
    })
    fmt_clara = workbook.add_format({'bg_color': '#F2F2F2'})
    fmt_escura = workbook.add_format({'bg_color': '#FFFFFF'})

    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, fmt_header)

    for row_num in range(1, len(df) + 1):
        fmt_linha = fmt_clara if row_num % 2 == 0 else fmt_escura
        for col_num in range(len(df.columns)):
            valor = df.iloc[row_num - 1, col_num]
            if isinstance(valor, str) and 'R$' in valor:
                try:
                    num = float(valor.replace('R$', '').replace(',', '.').strip())
                    worksheet.write(row_num, col_num, num, fmt_preco)
                except Exception:
                    worksheet.write(row_num, col_num, valor, fmt_linha)
            else:
                worksheet.write(row_num, col_num, valor, fmt_linha)


def _gerar_excel(df):
    """Gera bytes de um Excel formatado a partir de um DataFrame de resultados."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Preços', index=False)
        formatar_preco_excel(writer, 'Preços', df)
    return buffer.getvalue()


def _gerar_modelo_excel():
    """
    Gera bytes do arquivo Excel modelo para o usuário preencher.
    Cabeçalho azul escuro, CODIGO_BARRAS formatado como texto,
    larguras ajustadas e 3 linhas de exemplo com medicamentos fictícios.
    Usa openpyxl como engine.
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    df = pd.DataFrame({
        'CODIGO_BARRAS': ['7896181900018', '7891058001801', '7896261300012'],
        'DESCRICAO': [
            'Dipirona Sódica 500mg 20 comprimidos',
            'Amoxicilina 500mg 21 cápsulas',
            'Omeprazol 20mg 28 cápsulas',
        ],
        'CLASSIFICACAO': ['Analgésico', 'Antibiótico', 'Gastroprotetor'],
        'VALOR_ATUAL': [8.90, 24.50, 12.75],
    })
    df['CODIGO_BARRAS'] = df['CODIGO_BARRAS'].astype(str)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Modelo', index=False)
        ws = writer.sheets['Modelo']

        # Estilos do cabeçalho
        fill_azul = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        fonte_branca = Font(bold=True, color='FFFFFF')
        centralizado = Alignment(horizontal='center', vertical='center')

        # Aplicar formatação em cada célula do cabeçalho (linha 1)
        for col_idx in range(1, len(df.columns) + 1):
            cel = ws.cell(row=1, column=col_idx)
            cel.fill = fill_azul
            cel.font = fonte_branca
            cel.alignment = centralizado

        # CODIGO_BARRAS como texto para preservar zeros à esquerda
        for row_idx in range(2, len(df) + 2):
            cel = ws.cell(row=row_idx, column=1)
            cel.value = str(df.iloc[row_idx - 2]['CODIGO_BARRAS'])
            cel.number_format = '@'

        # VALOR_ATUAL com formato de moeda
        for row_idx in range(2, len(df) + 2):
            ws.cell(row=row_idx, column=4).number_format = 'R$ #,##0.00'

        # Larguras das colunas
        ws.column_dimensions['A'].width = 22  # CODIGO_BARRAS
        ws.column_dimensions['B'].width = 45  # DESCRICAO
        ws.column_dimensions['C'].width = 20  # CLASSIFICACAO
        ws.column_dimensions['D'].width = 16  # VALOR_ATUAL

    return buffer.getvalue()


def _colorir_linha(row, colunas_fontes):
    """Pinta a linha conforme a quantidade de preços encontrados."""
    encontrados = sum(
        1 for col in row.index
        if col in colunas_fontes and 'R$' in str(row[col])
    )
    total = len(colunas_fontes)
    if total == 0:
        return [''] * len(row)
    if encontrados == total:
        return ['background-color: lightgreen'] * len(row)
    elif encontrados > 0:
        return ['background-color: lightyellow'] * len(row)
    else:
        return ['background-color: lightcoral'] * len(row)


# ─────────────────────────────────────────────
# INICIALIZAR SESSION STATE
# ─────────────────────────────────────────────

def _init_session_state():
    """Garante que todas as variáveis de controle existam no session_state."""
    defaults = {
        'arquivo_carregado': False,
        'pesquisa_rodando': False,
        'pesquisa_concluida': False,
        'resultados': [],
        'df_input': None,
        'df_resultado': None,
        'arquivo_csv': None,
    }
    for chave, valor in defaults.items():
        if chave not in st.session_state:
            st.session_state[chave] = valor


# ─────────────────────────────────────────────
# INTERFACE PRINCIPAL
# ─────────────────────────────────────────────

def main():
    import os
    if not os.path.exists('logs'):
        os.makedirs('logs')
    if not os.path.exists('historico'):
        os.makedirs('historico')

    st.title("💊 Bot de Preços v2.0")

    _init_session_state()

    fontes_ativas = carregar_fontes()
    estrategia = carregar_estrategia()

    tab1, tab2, tab3, tab4 = st.tabs([
        "📤 Nova Pesquisa",
        "📊 Histórico",
        "🌐 Fontes",
        "⚙️ Logs e Configurações",
    ])

    # ═══════════════════════════════════════════
    # ABA 1 — NOVA PESQUISA
    # ═══════════════════════════════════════════
    with tab1:
        st.header("Nova Pesquisa de Preços")

        # ────────────────────────────────────────
        # BLOCO 1 — sempre visível: modelo
        # ────────────────────────────────────────
        st.subheader("📋 Modelo de Planilha")
        st.caption(
            "Sua planilha deve seguir o modelo abaixo. "
            "Baixe, preencha e importe aqui."
        )
        st.download_button(
            label="⬇ Baixar Modelo Excel",
            data=_gerar_modelo_excel(),
            file_name="modelo_produtos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.divider()

        # ────────────────────────────────────────
        # BLOCO 2 — sempre visível: upload
        # (nunca dentro de if/else — sempre renderizado)
        # ────────────────────────────────────────
        st.subheader("📂 Importar sua Planilha")
        uploaded_file = st.file_uploader(
            "Selecione o arquivo Excel ou CSV com seus produtos",
            type=["xlsx", "csv"],
        )

        # Lógica de estado — não renderiza widgets aqui, só atualiza session_state
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith(".xlsx"):
                    df_lido = pd.read_excel(uploaded_file)
                else:
                    df_lido = pd.read_csv(uploaded_file, sep=";", encoding="utf-8-sig")

                # Atualizar estado apenas na primeira carga (evita reprocessar a cada rerun)
                if not st.session_state.arquivo_carregado:
                    st.session_state.df_input = df_lido
                    st.session_state.arquivo_carregado = True
                    st.session_state.pesquisa_concluida = False
                    st.session_state.resultados = []
                    st.session_state.df_resultado = None
                    st.session_state.arquivo_csv = None

            except Exception as e:
                st.error(f"❌ Erro ao carregar arquivo: {e}")
                st.session_state.arquivo_carregado = False
                st.session_state.df_input = None

        else:
            # Arquivo removido pelo usuário — resetar só se pesquisa não foi concluída
            if st.session_state.arquivo_carregado and not st.session_state.pesquisa_concluida:
                st.session_state.arquivo_carregado = False
                st.session_state.pesquisa_rodando = False
                st.session_state.df_input = None
                st.session_state.resultados = []

        # ────────────────────────────────────────
        # BLOCO 3 — condicional: configs + botão
        # aparece somente após arquivo carregado
        # ────────────────────────────────────────
        if st.session_state.arquivo_carregado and st.session_state.df_input is not None:
            df_input = st.session_state.df_input
            colunas = list(df_input.columns)

            st.success(f"✅ Arquivo carregado: {len(df_input)} produtos")
            st.dataframe(df_input.head(5), use_container_width=True)
            st.info(f"📊 Colunas detectadas: {', '.join(colunas)}")

            # ── ESTADO: aguardando início ─────────────────────────
            if not st.session_state.pesquisa_rodando and not st.session_state.pesquisa_concluida:

                col_ean = st.selectbox(
                    "Coluna EAN/Código", colunas,
                    index=next((i for i, c in enumerate(colunas) if 'ean' in c.lower() or 'codigo' in c.lower()), 0),
                )
                col_nome = st.selectbox(
                    "Coluna Nome/Descrição", colunas,
                    index=next((i for i, c in enumerate(colunas) if 'desc' in c.lower() or 'nome' in c.lower()), 0),
                )

                st.subheader("⚙️ Configurações da Pesquisa")
                c1, c2 = st.columns(2)
                with c1:
                    lote = st.slider("Produtos por lote", 1, 10, 6)
                    inicio = st.number_input("Iniciar do produto", 1, 1000, 1)
                with c2:
                    espera_min = st.slider("Espera mínima (s)", 1, 10, 3)
                    espera_max = st.slider("Espera máxima (s)", 5, 20, 7)

                # ── MELHORIA 3: Fornecedores dinâmicos ───────────
                st.subheader("🌐 Fornecedores")
                fontes_disponiveis = [f for f in fontes_ativas if f.get("status") == "ativa"]

                if not fontes_disponiveis:
                    st.warning("Nenhuma fonte ativa. Configure na aba Fontes.")

                fontes_selecionadas = []
                for fonte in fontes_disponiveis:
                    key = f"tab1_fonte_{fonte['coluna_csv']}"
                    # Inicializar como True se ainda não existe no session_state
                    if key not in st.session_state:
                        st.session_state[key] = True
                    if st.checkbox(
                        f"✅ {fonte['nome']} (Método: {fonte['metodo']})",
                        key=key,
                    ):
                        fontes_selecionadas.append(fonte)

                if not fontes_selecionadas and fontes_disponiveis:
                    st.error("⚠ Selecione pelo menos um fornecedor para pesquisar.")

                # Botão iniciar
                btn_disabled = not fontes_selecionadas
                if st.button("▶ Iniciar Pesquisa", type="primary", disabled=btn_disabled):
                    st.session_state.pesquisa_rodando = True
                    st.session_state.resultados = []
                    configurar_logging()

                    # Montar lista do lote a processar
                    todos = []
                    for _, row in df_input.iterrows():
                        ean = str(row[col_ean]).strip().split(".")[0]
                        nome = str(row[col_nome]).strip() if col_nome in df_input.columns else ""
                        if ean:
                            todos.append({"ean": ean, "nome": nome})

                    lote_produtos = todos[int(inicio) - 1: int(inicio) - 1 + int(lote)]
                    total_lote = len(lote_produtos)
                    colunas_fontes = [f["coluna_csv"] for f in fontes_selecionadas]

                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    tabela_resultados = st.empty()

                    for idx, produto in enumerate(lote_produtos):
                        num = int(inicio) + idx
                        status_text.text(
                            f"🔍 Processando {num}/{int(inicio) - 1 + total_lote}: {produto['nome']}"
                        )

                        precos = buscar_produto(
                            produto["ean"], produto["nome"],
                            fontes_selecionadas, estrategia,
                        )

                        linha = {
                            "EAN": produto["ean"],
                            "Nome": produto["nome"],
                            "Data": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        }
                        for fonte in fontes_selecionadas:
                            coluna = fonte["coluna_csv"]
                            preco = precos.get(fonte["nome"], {}).get("preco", "Não encontrado")
                            linha[coluna] = preco
                            linha[f"Link_{coluna}"] = precos.get(fonte["nome"], {}).get("link", "")

                        st.session_state.resultados.append(linha)

                        df_atual = pd.DataFrame(st.session_state.resultados)
                        styled = df_atual.style.apply(
                            _colorir_linha, colunas_fontes=colunas_fontes, axis=1
                        )
                        tabela_resultados.dataframe(styled, use_container_width=True)
                        progress_bar.progress((idx + 1) / total_lote)

                        if idx < total_lote - 1:
                            time.sleep(random.uniform(espera_min, espera_max))

                    progress_bar.progress(1.0)
                    status_text.text("✅ Pesquisa concluída!")

                    df_final = pd.DataFrame(st.session_state.resultados)
                    arquivo_csv = salvar_csv_historico(df_final)

                    st.session_state.df_resultado = df_final
                    st.session_state.arquivo_csv = arquivo_csv
                    st.session_state.pesquisa_rodando = False
                    st.session_state.pesquisa_concluida = True

                    st.success(f"💾 Resultados salvos em: {arquivo_csv}")
                    st.download_button(
                        label="⬇ Baixar Excel Formatado",
                        data=_gerar_excel(df_final),
                        file_name=f"pesquisa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

            # ── ESTADO: pesquisa em progresso ─────────────────────
            elif st.session_state.pesquisa_rodando:
                st.info("⏳ Pesquisa em andamento...")
                st.button("⏳ Pesquisando...", disabled=True)

            # ── ESTADO: pesquisa concluída ─────────────────────────
            elif st.session_state.pesquisa_concluida:
                st.success("✅ Pesquisa concluída!")

                if st.session_state.df_resultado is not None:
                    df_resultado = st.session_state.df_resultado
                    colunas_fontes = [
                        col for col in df_resultado.columns
                        if not col.startswith("Link_") and col not in ["EAN", "Nome", "Data"]
                    ]
                    styled = df_resultado.style.apply(
                        _colorir_linha, colunas_fontes=colunas_fontes, axis=1
                    )
                    st.dataframe(styled, use_container_width=True)
                    st.download_button(
                        label="⬇ Baixar Excel Formatado",
                        data=_gerar_excel(df_resultado),
                        file_name=f"pesquisa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                if st.session_state.arquivo_csv:
                    st.info(f"📁 Arquivo CSV: {st.session_state.arquivo_csv}")

                if st.button("🔄 Nova Pesquisa"):
                    for chave in [
                        "arquivo_carregado", "pesquisa_rodando", "pesquisa_concluida",
                        "resultados", "df_input", "df_resultado", "arquivo_csv",
                    ]:
                        padrao = [] if chave == "resultados" else (
                            False if chave in ("arquivo_carregado", "pesquisa_rodando", "pesquisa_concluida")
                            else None
                        )
                        st.session_state[chave] = padrao
                    st.rerun()

    # ═══════════════════════════════════════════
    # ABA 2 — HISTÓRICO
    # ═══════════════════════════════════════════
    with tab2:
        st.header("📊 Histórico de Pesquisas")

        df_historico = carregar_historico()

        if df_historico.empty:
            st.info("📭 Nenhum histórico encontrado.")
        else:
            pesquisas = ["Todas combinadas"] + df_historico["arquivo_origem"].unique().tolist()
            pesquisa_selecionada = st.selectbox("Selecionar pesquisa", pesquisas)

            df_filtrado = (
                df_historico.copy() if pesquisa_selecionada == "Todas combinadas"
                else df_historico[df_historico["arquivo_origem"] == pesquisa_selecionada].copy()
            )

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                filtro_texto = st.text_input("🔍 Filtrar por nome/EAN")
            with c2:
                colunas_preco = [
                    col for col in df_filtrado.columns
                    if not col.startswith("Link_")
                    and col not in ["EAN", "Nome", "Data", "arquivo_origem"]
                ]
                filtro_site = st.selectbox("🏪 Filtrar por site", ["Todos"] + colunas_preco)
            with c3:
                filtro_resultado = st.selectbox(
                    "📊 Filtrar por resultado",
                    ["Todos", "Encontrados", "Não encontrados"],
                )
            with c4:
                if st.button("🧹 Limpar filtros"):
                    filtro_texto = ""
                    filtro_site = "Todos"
                    filtro_resultado = "Todos"

            if filtro_texto:
                df_filtrado = df_filtrado[
                    df_filtrado["Nome"].str.contains(filtro_texto, case=False, na=False) |
                    df_filtrado["EAN"].str.contains(filtro_texto, na=False)
                ]

            if filtro_site != "Todos":
                if filtro_resultado == "Encontrados":
                    df_filtrado = df_filtrado[
                        df_filtrado[filtro_site].notna() &
                        (df_filtrado[filtro_site] != "Não encontrado")
                    ]
                elif filtro_resultado == "Não encontrados":
                    df_filtrado = df_filtrado[df_filtrado[filtro_site] == "Não encontrado"]

            st.dataframe(df_filtrado.drop(columns=["arquivo_origem"]), use_container_width=True)

            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric("Total exibido", len(df_filtrado))

            medias = {}
            for col in colunas_preco:
                valores = pd.to_numeric(
                    df_filtrado[col].str.replace("R$", "").str.replace(",", ".").str.strip(),
                    errors="coerce",
                )
                media = valores.mean()
                if not pd.isna(media):
                    medias[col] = media

            with c2:
                if medias:
                    mais_barato = min(medias, key=medias.get)
                    st.metric("Mais barato (média)", f"{mais_barato}: R$ {medias[mais_barato]:.2f}")

            with c3:
                if colunas_preco:
                    todos_valores = pd.concat([
                        pd.to_numeric(
                            df_filtrado[col].str.replace("R$", "").str.replace(",", ".").str.strip(),
                            errors="coerce",
                        )
                        for col in colunas_preco
                    ])
                    media_geral = todos_valores.mean()
                    st.metric("Média geral", f"R$ {media_geral:.2f}" if not pd.isna(media_geral) else "N/A")

            if medias:
                fig = px.bar(
                    x=list(medias.keys()), y=list(medias.values()),
                    title="Média de Preços por Site",
                    labels={"x": "Site", "y": "Preço Médio (R$)"},
                )
                st.plotly_chart(fig, use_container_width=True)

            if st.button("⬇ Exportar Seleção em Excel"):
                df_exp = df_filtrado.drop(columns=["arquivo_origem"])
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
                    df_exp.to_excel(writer, sheet_name="Histórico", index=False)
                    formatar_preco_excel(writer, "Histórico", df_exp)
                st.download_button(
                    label="📥 Download Excel",
                    data=buf.getvalue(),
                    file_name=f"historico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ═══════════════════════════════════════════
    # ABA 3 — FONTES  (MELHORIA 2)
    # ═══════════════════════════════════════════
    with tab3:
        st.header("🌐 Gestão de Fontes")

        dados_fontes = _carregar_fontes_json()

        # ── SEÇÃO A: Fontes Cadastradas ───────────────────────────
        st.subheader("A — Fontes Cadastradas")
        st.caption("Ative ou pause cada fonte. Fontes inativas não aparecem na pesquisa.")

        fontes_todas = dados_fontes.get("fontes", [])

        if not fontes_todas:
            st.info("Nenhuma fonte cadastrada ainda.")
        else:
            # Cabeçalho da tabela
            h0, h1, h2, h3, h4 = st.columns([2, 3, 2, 1, 2])
            h0.markdown("**Nome**")
            h1.markdown("**URL**")
            h2.markdown("**Método**")
            h3.markdown("**Status**")
            h4.markdown("**Ações**")
            st.divider()

            for i, fonte in enumerate(fontes_todas):
                status = fonte.get("status", "pendente")
                c0, c1, c2, c3, c4 = st.columns([2, 3, 2, 1, 2])

                with c0:
                    if status == "inativa":
                        st.markdown(
                            f"<span style='color:#999;text-decoration:line-through'>"
                            f"{fonte['nome']}</span>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.write(fonte["nome"])

                with c1:
                    if status == "inativa":
                        st.markdown(
                            f"<span style='color:#999'>{fonte['url']}</span>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.write(fonte["url"])

                with c2:
                    st.write(fonte.get("metodo", "N/A"))

                with c3:
                    if status == "ativa":
                        st.success("Ativa")
                    elif status == "inativa":
                        st.error("Inativa")
                    else:
                        st.warning("Pendente")

                with c4:
                    btn1, btn2 = st.columns(2)
                    with btn1:
                        if st.button("✅ Ativar", key=f"ativar_{i}",
                                     disabled=(status == "ativa")):
                            _atualizar_status_fonte(fonte["nome"], "ativa")
                            st.rerun()
                    with btn2:
                        if st.button("⏸ Pausar", key=f"pausar_{i}",
                                     disabled=(status == "inativa")):
                            _atualizar_status_fonte(fonte["nome"], "inativa")
                            st.rerun()

        st.divider()

        # ── SEÇÃO B: Adicionar Nova Fonte ─────────────────────────
        st.subheader("B — Adicionar Nova Fonte")

        with st.form("nova_fonte"):
            nome_novo = st.text_input("Nome da farmácia", placeholder="Ex: Ultrafarma")
            url_novo = st.text_input(
                "URL do site",
                placeholder="https://www.exemplo.com.br",
            )
            coluna_novo = st.text_input(
                "Nome da coluna no CSV (maiúsculo, sem espaços)",
                placeholder="Ex: ULTRAFARMA, PANVEL, DROGA_RAIA",
            )

            if st.form_submit_button("➕ Adicionar Fonte"):
                erros = []
                if not nome_novo.strip():
                    erros.append("Informe o nome da farmácia.")
                if not url_novo.strip().startswith("https://"):
                    erros.append("A URL deve começar com https://")
                col_limpa = coluna_novo.strip().upper().replace(" ", "_")
                if not col_limpa:
                    erros.append("Informe o nome da coluna CSV.")
                elif coluna_novo.strip() != col_limpa:
                    erros.append(
                        "Nome da coluna deve ser maiúsculo e sem espaços "
                        "(ex: ULTRAFARMA, DROGA_RAIA)."
                    )

                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    dados_atual = _carregar_fontes_json()
                    nova = {
                        "nome": nome_novo.strip(),
                        "url": url_novo.strip(),
                        "status": "pendente",
                        "metodo": "PLAYWRIGHT",
                        "coluna_csv": col_limpa,
                        "adicionada_em": datetime.now().strftime("%Y-%m-%d"),
                    }
                    dados_atual["pendentes"].append(nova)
                    salvar_fontes(dados_atual["fontes"], dados_atual["pendentes"])
                    st.warning(
                        "⚠ Fonte salva como pendente. Esta fonte ainda não será pesquisada "
                        "automaticamente. O desenvolvedor precisa mapear os seletores deste site "
                        "primeiro. Anote a URL e peça a atualização quando necessário."
                    )
                    st.rerun()

        # Lista de pendentes
        dados_atual = _carregar_fontes_json()
        if dados_atual.get("pendentes"):
            st.markdown("**⏳ Fontes Pendentes**")
            for i, pendente in enumerate(dados_atual["pendentes"]):
                pc1, pc2, pc3 = st.columns([3, 3, 1])
                with pc1:
                    st.write(pendente["nome"])
                with pc2:
                    st.write(pendente["url"])
                with pc3:
                    if st.button("🗑 Remover", key=f"remover_pendente_{i}"):
                        dados_atual["pendentes"].pop(i)
                        salvar_fontes(dados_atual["fontes"], dados_atual["pendentes"])
                        st.rerun()

        st.divider()

        # ── SEÇÃO C: Selecionar Fontes para a Pesquisa ───────────
        st.subheader("C — Selecionar Fontes para a Pesquisa")
        st.caption(
            "As fontes selecionadas aqui serão usadas em todas as "
            "próximas pesquisas desta sessão."
        )

        fontes_ativas_tab3 = [
            f for f in dados_fontes.get("fontes", [])
            if f.get("status") == "ativa"
        ]

        if not fontes_ativas_tab3:
            st.info("Nenhuma fonte ativa. Ative fontes na Seção A acima.")
        else:
            sb1, sb2 = st.columns(2)
            with sb1:
                if st.button("Selecionar todas", key="sel_todas"):
                    for f in fontes_ativas_tab3:
                        # Atualiza tanto os checkboxes do Tab3 quanto do Tab1
                        st.session_state[f"tab3_fonte_{f['coluna_csv']}"] = True
                        st.session_state[f"tab1_fonte_{f['coluna_csv']}"] = True
                    st.rerun()
            with sb2:
                if st.button("Desmarcar todas", key="des_todas"):
                    for f in fontes_ativas_tab3:
                        st.session_state[f"tab3_fonte_{f['coluna_csv']}"] = False
                        st.session_state[f"tab1_fonte_{f['coluna_csv']}"] = False
                    st.rerun()

            for fonte in fontes_ativas_tab3:
                tab3_key = f"tab3_fonte_{fonte['coluna_csv']}"
                # Inicializar com True se ainda não existe
                if tab3_key not in st.session_state:
                    st.session_state[tab3_key] = True
                st.checkbox(
                    f"✅ {fonte['nome']} (Método: {fonte['metodo']})",
                    key=tab3_key,
                    on_change=_sync_fonte_para_tab1,
                    args=(fonte["coluna_csv"],),
                )

    # ═══════════════════════════════════════════
    # ABA 4 — LOGS E CONFIGURAÇÕES
    # ═══════════════════════════════════════════
    with tab4:
        st.header("⚙️ Logs e Configurações")

        st.subheader("📋 Logs de Execução")

        if os.path.exists("logs"):
            arquivos_log = sorted(
                [f for f in os.listdir("logs") if f.endswith(".txt")],
                reverse=True,
            )
            if arquivos_log:
                log_selecionado = st.selectbox("Selecionar log", arquivos_log)

                c1, c2 = st.columns(2)
                with c1:
                    filtro_nivel = st.selectbox(
                        "Filtrar nível", ["Todos", "INFO", "WARNING", "ERROR"]
                    )
                with c2:
                    st.write("")

                try:
                    with open(f"logs/{log_selecionado}", "r", encoding="utf-8") as f:
                        conteudo = f.read()

                    if filtro_nivel != "Todos":
                        linhas_filtradas = [
                            l for l in conteudo.split("\n")
                            if f"[{filtro_nivel}]" in l
                        ]
                        conteudo_filtrado = "\n".join(linhas_filtradas)
                    else:
                        conteudo_filtrado = conteudo

                    st.code(conteudo_filtrado, language="text")

                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Total de linhas", len(conteudo_filtrado.split("\n")))
                    c2.metric("Erros", conteudo_filtrado.count("[ERROR]"))
                    c3.metric("Avisos", conteudo_filtrado.count("[WARNING]"))
                    c4.metric("Sucessos", conteudo_filtrado.count("[INFO]"))

                    if st.button("🗑 Limpar logs antigos (manter últimos 10)"):
                        if len(arquivos_log) > 10:
                            for arquivo in arquivos_log[10:]:
                                os.remove(f"logs/{arquivo}")
                            st.success("Logs antigos removidos!")
                            st.rerun()
                        else:
                            st.info("Não há logs antigos para remover.")

                except Exception as e:
                    st.error(f"Erro ao ler log: {e}")
            else:
                st.info("Nenhum log encontrado.")
        else:
            st.info("Diretório de logs não existe.")

        st.subheader("🔧 Configurações do Sistema")

        try:
            import config
            config_atual = {
                "LOTE": config.LOTE,
                "ARQUIVO_ENTRADA": config.ARQUIVO_ENTRADA,
                "ARQUIVO_SAIDA": config.ARQUIVO_SAIDA,
                "TEMPO_ESPERA_MIN": config.TEMPO_ESPERA_MIN,
                "TEMPO_ESPERA_MAX": config.TEMPO_ESPERA_MAX,
                "INICIAR_DO_PRODUTO": config.INICIAR_DO_PRODUTO,
                "VISIVEL": config.VISIVEL,
            }
        except Exception:
            config_atual = {}

        with st.form("config_form"):
            lote_novo = st.number_input("Produtos por lote", 1, 20, config_atual.get("LOTE", 6))
            entrada_novo = st.text_input("Arquivo entrada", config_atual.get("ARQUIVO_ENTRADA", "produtos.xlsx"))
            saida_novo = st.text_input("Arquivo saída", config_atual.get("ARQUIVO_SAIDA", "resultado.csv"))
            espera_min_novo = st.number_input("Espera min (s)", 1, 10, config_atual.get("TEMPO_ESPERA_MIN", 3))
            espera_max_novo = st.number_input("Espera max (s)", 5, 30, config_atual.get("TEMPO_ESPERA_MAX", 7))
            iniciar_novo = st.number_input("Iniciar do produto", 1, 1000, config_atual.get("INICIAR_DO_PRODUTO", 1))
            visivel_novo = st.checkbox("Navegador visível", config_atual.get("VISIVEL", False))

            salvar_clicado = st.form_submit_button("💾 Salvar Configurações")

        # ── fora do form: processar submit e botão auxiliar ──────
        if salvar_clicado:
            config_texto = (
                "# Configuracoes do Bot de Pesquisa de Precos\n\n"
                f"LOTE = {lote_novo}\n\n"
                f'ARQUIVO_ENTRADA = "{entrada_novo}"\n\n'
                f'ARQUIVO_SAIDA = "{saida_novo}"\n\n'
                f"TEMPO_ESPERA_MIN = {espera_min_novo}\n\n"
                f"TEMPO_ESPERA_MAX = {espera_max_novo}\n\n"
                f"INICIAR_DO_PRODUTO = {iniciar_novo}\n\n"
                f"VISIVEL = {str(visivel_novo).lower()}\n\n"
                "HEADERS = {\n"
                '    "User-Agent": (\n'
                '        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "\n'
                '        "AppleWebKit/537.36 (KHTML, like Gecko) "\n'
                '        "Chrome/124.0.0.0 Safari/537.36"\n'
                "    ),\n"
                '    "Accept-Language": "pt-BR,pt;q=0.9",\n'
                '    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",\n'
                "}\n\n"
                "SITES = []\n"
            )
            with open("config.py", "w", encoding="utf-8") as f:
                f.write(config_texto)
            st.success("✅ Configurações salvas!")

        if st.button("🗑 Apagar estrategia.json"):
            if os.path.exists("estrategia.json"):
                os.remove("estrategia.json")
                st.success("estrategia.json removido!")
            else:
                st.info("Arquivo não existe.")


if __name__ == "__main__":
    main()
